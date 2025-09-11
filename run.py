import click
import src.dataOps as dataOps
import src.writer as writer
import logging
import os
from tqdm import tqdm
from pathlib import Path
import shutil
import re
from datetime import datetime, date, time
from dateutil.parser import parse
import pandas as pd
from typing import List, Tuple, Iterable
import re
from datetime import datetime, date, time
from dateutil.parser import parse
from openpyxl import load_workbook


logger = logging.getLogger(__name__)

# ДОБАВЛЯЕМ настройки ширины терминала
CONTEXT_SETTINGS = {"max_content_width": shutil.get_terminal_size().columns}


@click.group(context_settings=CONTEXT_SETTINGS)
def main():
    logging.basicConfig(filename="log.txt", level=logging.INFO)
    pass


@main.command("data-operations")
@click.argument("path", type=click.Path(exists=True, path_type=Path))
@click.option(
    "--format",
    type=click.Choice(["cd", "xg"]),
    default="cd",
    show_default=True,
    help=(
        '"xg" — сырые данные из иксолоджи (CSV);\n'
        '"cd" — экспорт из ЦУСАД (первый лист Excel).'
    ),
)
@click.option(
    "-o",
    "--out-dir",
    "out_base",  # <= сразу в out_base
    type=click.Path(file_okay=False, dir_okay=True, path_type=Path),
    default=None,
    help="Куда сохранять результат. Если не указано — сохраняем рядом с исходником.",
)
@click.option(
    "--overwrite/--no-overwrite",
    default=False,
    show_default=True,
    help="Перезаписывать существующие файлы назначения.",
)
@click.option(
    "--suffix",
    default="_processed",
    show_default=True,
    help="Суффикс имени файла, когда --out-dir не задан.",
)
def data_operations(
    path: Path, format: str, out_base: Path | None, overwrite: bool, suffix: str
):
    """
    Обрабатывает файл/директорию PATH рекурсивно.
    Сохраняет структуру подпапок при использовании --out-dir.
    """
    # дальше используйте out_base внутри своей обработки:
    #   - если out_base задан: dst = out_base / src.relative_to(path)
    #   - иначе: dst = src.with_name(src.stem + suffix + src.suffix)
    #   - учитывайте overwrite при записи
    #
    # ваш существующий код обработки здесь ⤵
    ...
    extensions = {".xls", ".xlsx"} if (format == "cd") else {".csv"}
    path_object = Path(path)

    def is_datetime(value) -> bool:
        if isinstance(value, (datetime, date)):
            return True
        try:
            parse(str(value), fuzzy=False)
            return True
        except Exception:
            return False

    def process_file(src_path: Path):
        if src_path.name.startswith("~$"):
            return

        if format == "xg":
            df = dataOps.process_csv_dataframe(src_path)
            # пути сохранения
            if out_base:
                dst_path = out_base / src_path.relative_to(path_object)
            else:
                dst_path = src_path.with_name(src_path.stem + "_processed.csv")
            dst_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(dst_path, index=False, encoding="utf-8-sig")

        elif format == "cd":
            print(f"Обработка {src_path}")
            initial_df, predicted = dataOps.process_xslx_dataframe(src_path)

            # КУДА ПИСАТЬ: формируем ПУТЬ НОВОГО ФАЙЛА
            if out_base:
                dst_path = out_base / src_path.relative_to(path_object)
            else:
                dst_path = src_path.with_name(
                    src_path.stem + "_processed" + src_path.suffix
                )

            dst_path.parent.mkdir(parents=True, exist_ok=True)

            # write_to_predict_sheet САМ СОЗДАЁТ НОВЫЙ ФАЙЛ ПО ЭТОМУ ПУТИ
            print(f"Запись в новый файл {dst_path}")
            writer.write_to_predict_sheet(dst_path, initial_df, predicted)

    # собрать файлы рекурсивно
    files = []
    if path_object.is_dir():
        files = [
            f
            for f in path_object.rglob("*")
            if f.is_file() and f.suffix.lower() in extensions
        ]
    elif path_object.is_file() and path_object.suffix.lower() in extensions:
        files = [path_object]

    for filePath in tqdm(files, desc="Обработка файлов"):
        try:
            process_file(filePath)
        except BaseException as e:
            print(f"Ошибка при обработке {filePath}: {e}")


def _read_timestamp_col(path: Path) -> pd.Series:
    """
    Чтение первого столбца как меток времени.
    Формат таблицы: 1-я строка — категории ТС; 2-я — 'Итого/Прямое/Обратное';
    3-я — пустая; с 4-й — данные. Первый столбец — timestamp.
    """
    try:
        # двухуровневый заголовок + пропускаем 3-ю (пустую) строку
        df = pd.read_excel(
            path, header=[0, 1], skiprows=[2], sheet_name="Объединенные данные"
        )
    except Exception:
        # на всякий случай — более общий парсинг
        df = pd.read_excel(path)

    if df.shape[1] == 0:
        raise ValueError("В файле нет столбцов")

    ts = pd.to_datetime(df.iloc[:, 0], errors="coerce")
    ts = ts.dropna().sort_values().reset_index(drop=True)
    if ts.empty:
        raise ValueError("Не удалось распарсить столбец timestamp")
    return ts


def _missing_ranges_hourly(ts: pd.Series) -> List[Tuple[pd.Timestamp, pd.Timestamp]]:
    """
    Возвращает список пропущенных диапазонов (шаг 1 час) от первого до последнего ts.
    Если пропуск одиночный — start == end.
    """
    start, end = ts.iloc[0], ts.iloc[-1]
    expected = pd.date_range(start=start, end=end, freq="H")
    existing = pd.DatetimeIndex(ts.unique())
    missing = expected.difference(existing)
    if len(missing) == 0:
        return []

    ranges: List[Tuple[pd.Timestamp, pd.Timestamp]] = []
    run_start = missing[0]
    prev = missing[0]
    for t in missing[1:]:
        if (t - prev) == pd.Timedelta(hours=1):
            prev = t
        else:
            ranges.append((run_start, prev))
            run_start = prev = t
    ranges.append((run_start, prev))
    return ranges


def _count_missing_in_ranges(
    ranges: Iterable[Tuple[pd.Timestamp, pd.Timestamp]],
) -> int:
    total = 0
    for a, b in ranges:
        total += int((b - a) / pd.Timedelta(hours=1)) + 1
    return total


def _fmt_ts(t: pd.Timestamp) -> str:
    return t.strftime("%Y-%m-%d %H:%M:%S")


@main.command("check-missing", context_settings=CONTEXT_SETTINGS)
@click.argument(
    "root",
    type=click.Path(exists=True, file_okay=False, path_type=Path),
)
@click.option(
    "--ext",
    multiple=True,
    default=[".xlsx", ".xls"],
    show_default=True,
    help="Какие расширения файлов просматривать (можно несколько раз).",
)
@click.option(
    "--out",
    type=click.Path(path_type=Path),
    default=None,
    help="Путь к CSV-отчёту. Если не задан — только вывод в консоль.",
)
def check_missing(root: Path, ext: tuple[str, ...], out: Path | None):
    """
    Проверяет, что метки времени (1-й столбец) во ВСЕХ файлах и подпапках ROOT
    идут каждые 60 минут от первой до последней. Выводит интервалы пропусков.
    """
    exts = {e.lower() for e in ext}
    files = sorted(
        [p for p in Path(root).rglob("*") if p.is_file() and p.suffix.lower() in exts]
    )

    if not files:
        click.echo("Файлы не найдены.", err=True)
        return

    rows_for_csv = []
    total_missing = 0

    for f in files:
        if f.name.startswith("~$"):  # пропускаем временные файлы Excel
            continue

        rel = f.relative_to(root)
        try:
            ts = _read_timestamp_col(f)
            ranges = _missing_ranges_hourly(ts)
            first_ts, last_ts = ts.iloc[0], ts.iloc[-1]
            missing_n = _count_missing_in_ranges(ranges)
            total_missing += missing_n

            if missing_n == 0:
                click.echo(
                    f"✅ {rel}: пропусков нет ({_fmt_ts(first_ts)} → {_fmt_ts(last_ts)})"
                )
                ranges_str = ""
            else:
                click.echo(
                    f"⚠️ {rel}: пропущено {missing_n} отметок "
                    f"({_fmt_ts(first_ts)} → {_fmt_ts(last_ts)})"
                )
                for a, b in ranges:
                    if a == b:
                        click.echo(f"   • {_fmt_ts(a)}")
                    else:
                        hours = int((b - a) / pd.Timedelta(hours=1)) + 1
                        click.echo(f"   • {_fmt_ts(a)} — {_fmt_ts(b)}  ({hours} ч.)")
                ranges_str = " | ".join(
                    [
                        f"{_fmt_ts(a)}" if a == b else f"{_fmt_ts(a)} — {_fmt_ts(b)}"
                        for a, b in ranges
                    ]
                )

            rows_for_csv.append(
                {
                    "file": str(rel),
                    "first_timestamp": _fmt_ts(first_ts),
                    "last_timestamp": _fmt_ts(last_ts),
                    "missing_count": missing_n,
                    "missing_ranges": ranges_str,
                }
            )

        except Exception as e:
            click.echo(f"❌ {rel}: ошибка чтения — {e}", err=True)
            rows_for_csv.append(
                {
                    "file": str(rel),
                    "first_timestamp": "",
                    "last_timestamp": "",
                    "missing_count": "",
                    "missing_ranges": f"ERROR: {e}",
                }
            )

    if out is not None:
        df_out = pd.DataFrame(rows_for_csv)
        out.parent.mkdir(parents=True, exist_ok=True)
        df_out.to_csv(out, index=False, encoding="utf-8-sig")
        click.echo(f"\nСводный отчёт сохранён: {out}")

    click.echo(f"\nИтого пропущенных отметок по всем файлам: {total_missing}")


@main.command("copy-missing", context_settings=CONTEXT_SETTINGS)
@click.argument("report", type=click.Path(exists=True, dir_okay=False, path_type=Path))
@click.argument("root", type=click.Path(exists=True, file_okay=False, path_type=Path))
@click.option(
    "-o",
    "--out-dir",
    type=click.Path(file_okay=False, dir_okay=True, path_type=Path),
    required=True,
    help="Куда класть копии. Структура папок сохранится относительно ROOT.",
)
@click.option(
    "--include-ok/--only-missing",
    default=False,
    show_default=True,
    help="Копировать все файлы из отчёта, а не только с пропусками.",
)
@click.option(
    "--include-errors/--skip-errors",
    default=False,
    show_default=True,
    help="Также копировать файлы, по которым в отчёте была ошибка чтения.",
)
def copy_missing(
    report: Path, root: Path, out_dir: Path, include_ok: bool, include_errors: bool
):
    """
    Читает CSV-отчёт из 'check-missing' и копирует указанные там файлы
    (по умолчанию — только те, у которых есть пропуски) в OUT_DIR,
    сохраняя относительную структуру относительно ROOT.
    """
    try:
        df = pd.read_csv(report)
    except Exception as e:
        click.echo(f"Не удалось прочитать отчёт {report}: {e}", err=True)
        return

    if "file" not in df.columns:
        click.echo("В отчёте нет столбца 'file'.", err=True)
        return

    # Приводим missing_count к числу
    mc = pd.to_numeric(
        df.get("missing_count", pd.Series([None] * len(df))), errors="coerce"
    )
    df["missing_count_num"] = mc

    # Фильтрация
    if include_ok:
        mask = df["file"].notna()
    else:
        mask = df["missing_count_num"].fillna(0) > 0

    if include_errors:
        err_mask = df["missing_count_num"].isna() & df.get("missing_ranges", "").astype(
            str
        ).str.startswith("ERROR")
        mask = mask | err_mask

    selected = df.loc[mask].copy()
    if selected.empty:
        click.echo("Подходящих файлов по отчёту не найдено.", err=True)
        return

    root_res = root.resolve()
    copied = 0
    missing = 0
    for _, row in tqdm(
        selected.iterrows(), total=selected.shape[0], desc="Копирование"
    ):
        rel = Path(str(row["file"]))
        src = (root_res / rel).resolve()

        # Защита от выходов из ROOT
        if src != root_res and root_res not in src.parents:
            click.echo(
                f"⚠️ Строка отчёта указывает путь вне ROOT, пропуск: {rel}", err=True
            )
            missing += 1
            continue

        if not src.exists():
            click.echo(f"⚠️ Не найден исходный файл, пропуск: {src}", err=True)
            missing += 1
            continue

        dst = (out_dir / rel).resolve()
        dst.parent.mkdir(parents=True, exist_ok=True)
        try:
            shutil.copy2(src, dst)
            copied += 1
        except Exception as e:
            click.echo(f"❌ Ошибка копирования {src} → {dst}: {e}", err=True)

    click.echo(
        f"\nГотово. Скопировано файлов: {copied}. Не найдено/пропущено: {missing}."
    )


@main.command("split-from-header", context_settings=CONTEXT_SETTINGS)
@click.argument(
    "input_path",
    type=click.Path(exists=True, file_okay=True, dir_okay=True, path_type=Path),
)
@click.option(
    "--header",
    type=click.Path(exists=True, dir_okay=False, path_type=Path),
    default=Path("header.xlsx"),
    show_default=True,
    help="Шаблон/хедер Excel, копия которого будет основой для каждого блока.",
)
@click.option(
    "-o",
    "--out-dir",
    "output_base",
    type=click.Path(file_okay=False, dir_okay=True, path_type=Path),
    default=Path("output_data"),
    show_default=True,
    help="Базовая папка для выходных файлов.",
)
def split_from_header(input_path: Path, header: Path, output_base: Path):
    """
    Режет первый лист каждого XLSX в папке ИЛИ одиночный XLSX-файл на блоки по строкам вида 'км N+M'
    и пишет каждый блок в отдельный .xlsx поверх header (логика не изменена).
    """
    # === НАСТРОЙКИ ===
    HEADER_XLSX = header.resolve()
    OUTPUT_BASE = output_base.resolve()
    OUTPUT_BASE.mkdir(parents=True, exist_ok=True)

    # === РЕГУЛЯРКИ ===
    km_re = re.compile(r"км \d+\+\d+", re.IGNORECASE)

    # === ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ (логика исходная) ===
    def safe_filename(s: str) -> str:
        s = s.strip()
        s = s.replace('"', "")
        s = s.replace(" ", "_")
        s = s.replace("/", "_")
        s = re.sub(r'[\\:*?"<>|]', "", s)
        s = re.sub(r"_+", "_", s)
        return s

    def is_datetime(value) -> bool:
        if isinstance(value, (datetime, date, time)):
            return True
        try:
            parse(str(value), fuzzy=False)
            return True
        except (ValueError, TypeError):
            return False

    def is_valid_highway_line(s) -> bool:
        s_str = (str(s) if s is not None else "").strip()
        if not s_str or s_str.lower() in ["итого", "среднее", "%"]:
            return False
        if km_re.search(s_str):
            return False
        if is_datetime(s):
            return False
        return True

    def should_process_file(p: Path) -> bool:
        """Фильтр входных файлов: .xlsx, не временный и не сам header.xlsx."""
        if not p.is_file():
            return False
        if p.suffix.lower() != ".xlsx":
            return False
        if p.name.startswith("~$"):
            return False
        try:
            return p.resolve() != HEADER_XLSX
        except Exception:
            return True  # на всякий случай, если resolve() упадёт

    def process_xlsx(input_xlsx_path: Path) -> None:
        """Режем первый лист входного XLSX на блоки и сохраняем каждый блок в .xlsx с header сверху."""
        current_highway = None
        current_km = None
        out_wb = None  # активная выходная книга (копия HEADER_XLSX)
        out_ws = None  # лист, куда пишем
        out_path: Path | None = None  # путь к текущему выходному файлу
        prev_row = None

        src_wb = load_workbook(input_xlsx_path, data_only=True, read_only=True)
        try:
            src_ws = src_wb.worksheets[0]

            for row in src_ws.iter_rows(values_only=True):
                if not row or (row[0] is None) or (str(row[0]).strip() == ""):
                    prev_row = row
                    continue

                first_cell_raw = row[0]
                first_cell = str(first_cell_raw).strip()

                # Обнаружение новой КМО
                if km_re.search(first_cell):
                    # Закрыть предыдущий выходной файл, если был
                    if out_wb and out_path:
                        out_wb.save(out_path)
                        out_wb.close()
                        out_wb, out_ws, out_path = None, None, None

                    current_km = safe_filename(first_cell)

                    # Попытка взять трассу из предыдущей строки
                    prev_val = prev_row[0] if prev_row and len(prev_row) > 0 else ""
                    if is_valid_highway_line(prev_val):
                        current_highway = safe_filename(str(prev_val))

                    if not current_highway:
                        print(
                            f"[!] Пропущен блок КМО: {current_km} (трасса не определена)"
                        )
                        prev_row = row
                        continue

                    # Создаём выходной XLSX из шаблона header.xlsx
                    highway_dir = OUTPUT_BASE / current_highway
                    highway_dir.mkdir(parents=True, exist_ok=True)
                    out_path = highway_dir / f"{current_km}.xlsx"

                    # Загружаем шаблон без read_only, чтобы дописывать
                    out_wb = load_workbook(HEADER_XLSX)
                    out_ws = (
                        out_wb.active
                    )  # дописываем данные блока в конец этого листа
                    prev_row = row
                    continue

                # Строка данных блока
                if out_ws and is_datetime(first_cell_raw):
                    out_ws.append(list(row))

                prev_row = row

            # Сохранить последний открытый файл
            if out_wb and out_path:
                out_wb.save(out_path)
                out_wb.close()

        finally:
            try:
                src_wb.close()
            except Exception:
                pass

    # === СБОР СПИСКА ВХОДНЫХ ФАЙЛОВ ===
    if input_path.is_dir():
        candidates = [p for p in input_path.iterdir() if should_process_file(p)]
    else:
        candidates = [input_path] if should_process_file(input_path) else []

    if not candidates:
        click.echo(f"Не найдено подходящих .xlsx для обработки в: {input_path}")
        return

    # === ОБРАБОТКА ВСЕХ ФАЙЛОВ ===
    for in_path in sorted(candidates):
        click.echo(f"▶ Обработка: {in_path}")
        try:
            process_xlsx(in_path)
            click.echo(f"✅ Готово: {in_path}")
        except Exception as e:
            click.echo(f"❌ Ошибка при обработке {in_path}: {e}")


@main.command("compare-structure", context_settings=CONTEXT_SETTINGS)
@click.argument(
    "path", type=click.Path(exists=True, file_okay=False, dir_okay=True, path_type=Path)
)
@click.argument(
    "path2",
    type=click.Path(exists=True, file_okay=False, dir_okay=True, path_type=Path),
)
@click.option(
    "--prefix",
    default="Обработано__",
    show_default=True,
    help="Префикс в именах файлов, который нужно игнорировать при сравнении. "
    "После снятия также убираются ведущие '_'/' ' после префикса.",
)
@click.option(
    "--prefix-side",
    type=click.Choice(["path", "path2", "both"]),
    default="path",
    show_default=True,
    help="На какой стороне снимать префикс из имён файлов (PATH, PATH2 или обеих). "
    "Для вашего случая укажите: path2.",
)
@click.option(
    "--ext",
    multiple=True,
    default=[".xlsx", ".xls"],
    show_default=True,
    help="Какие расширения сравнивать (можно указать несколько раз).",
)
@click.option(
    "--equate-excel-ext/--strict-ext",
    default=True,
    show_default=True,
    help="Считать .xls и .xlsx эквивалентными (упрощает сравнение).",
)
@click.option(
    "--show-extra/--no-show-extra",
    default=False,
    show_default=True,
    help="Показать также файлы, которые есть в PATH2, но отсутствуют в PATH (после нормализации).",
)
def compare_structure(
    path: Path,
    path2: Path,
    prefix: str,
    prefix_side: str,
    ext: tuple[str, ...],
    equate_excel_ext: bool,
    show_extra: bool,
):
    """
    Проверяет, что все файлы и папки в PATH соответствуют такой же структуре в PATH2.
    Префикс может быть на стороне PATH, PATH2 или на обеих — задаётся --prefix-side.
    При включённом --equate-excel-ext расширения .xls и .xlsx считаются одинаковыми.
    """
    import unicodedata

    exts = {e.lower() for e in ext}

    def u_norm(s: str) -> str:
        # Нормализация Unicode, чтобы убрать скрытые расхождения
        return unicodedata.normalize("NFC", s)

    def strip_prefix(name: str, pref: str) -> str:
        name = u_norm(name)
        pref = u_norm(pref)
        if name.startswith(pref):
            rest = name[len(pref) :]
            # убираем ведущие подчёркивания/пробелы/дефисы после префикса
            rest = re.sub(r"^[_\s-]+", "", rest)
            return rest
        return name

    def normalize_basename(basename: str, apply_prefix: bool) -> str:
        """Снимает префикс и при необходимости нормализует расширение."""
        name = strip_prefix(basename, prefix) if apply_prefix else basename
        name = u_norm(name)

        # Выравнивание расширений: .xls == .xlsx (только для сравнения)
        if equate_excel_ext:
            suff = Path(name).suffix.lower()
            if suff in {".xls", ".xlsx"}:
                name = Path(name).with_suffix(".xlsx").name
        return name

    def normalize_rel(rel: Path, side_tag: str) -> Path:
        """Возвращает относительный путь с нормализованным basename (только файл)."""
        apply = prefix_side in (side_tag, "both")
        norm_name = normalize_basename(rel.name, apply)
        return rel.with_name(norm_name)

    # --- ЛЕВАЯ сторона (PATH) ---
    normalized_from: set[Path] = set()
    collisions_left: dict[Path, list[Path]] = {}

    for p in path.rglob("*"):
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() not in exts:
            continue

        rel = p.relative_to(path)
        rel_norm = normalize_rel(rel, "path")
        if rel_norm in normalized_from and prefix_side in ("path", "both"):
            collisions_left.setdefault(rel_norm, []).append(rel)
        normalized_from.add(rel_norm)

    # --- ПРАВАЯ сторона (PATH2) ---
    normalized_to: set[Path] = set()
    collisions_right: dict[Path, list[Path]] = {}

    for q in path2.rglob("*"):
        if not q.is_file():
            continue
        if q.name.startswith("~$"):
            continue
        if q.suffix.lower() not in exts:
            continue

        rel = q.relative_to(path2)
        rel_norm = normalize_rel(rel, "path2")
        if rel_norm in normalized_to and prefix_side in ("path2", "both"):
            collisions_right.setdefault(rel_norm, []).append(rel)
        normalized_to.add(rel_norm)

    # --- Директории, ожидаемые в PATH2 (по структуре слева) ---
    expected_dirs = {p.parent for p in normalized_from}
    missing_dirs = sorted([d for d in expected_dirs if not (path2 / d).exists()])

    # --- Сравнение ---
    missing_files = sorted([p for p in normalized_from if p not in normalized_to])
    extra_files = (
        sorted([p for p in normalized_to if p not in normalized_from])
        if show_extra
        else []
    )

    # --- Вывод ---
    click.echo(
        f"Сравнение:\n"
        f"  FROM: {path}\n"
        f"  TO  : {path2}\n"
        f"  Префикс: '{prefix}'  (side: {prefix_side})\n"
        f"  Эквивалентность .xls/.xlsx: {'ON' if equate_excel_ext else 'OFF'}\n"
    )

    if collisions_left:
        click.echo("⚠️ Коллизии после снятия префикса на стороне PATH:")
        for rel_norm, originals in collisions_left.items():
            lst = "; ".join(str(o) for o in originals)
            click.echo(f"   • {rel_norm} <= {lst}")
        click.echo("")

    if collisions_right:
        click.echo("⚠️ Коллизии после снятия префикса на стороне PATH2:")
        for rel_norm, originals in collisions_right.items():
            lst = "; ".join(str(o) for o in originals)
            click.echo(f"   • {rel_norm} <= {lst}")
        click.echo("")

    if missing_dirs:
        click.echo("📁 Отсутствующие директории в PATH2:")
        for d in missing_dirs:
            click.echo(f"   • {d}")
        click.echo("")

    if missing_files:
        click.echo("📄 Отсутствующие файлы в PATH2:")
        for f in missing_files:
            click.echo(f"   • {f}")
    else:
        click.echo(
            "✅ Все ожидаемые файлы из PATH присутствуют в PATH2 (с учётом нормализации)."
        )

    if show_extra:
        click.echo("")
        if extra_files:
            click.echo("➕ Файлы, которые есть в PATH2, но отсутствуют в PATH:")
            for f in extra_files:
                click.echo(f"   • {f}")
        else:
            click.echo("➕ В PATH2 нет «лишних» файлов относительно PATH.")

    click.echo("\nИтоги:")
    click.echo(f"  Ожидалось (нормализовано): {len(normalized_from)}")
    click.echo(f"  Найдено в PATH2: {len(normalized_from) - len(missing_files)}")
    click.echo(f"  Отсутствует: {len(missing_files)}")
    if show_extra:
        click.echo(f"  «Лишних» в PATH2: {len(extra_files)}")


@main.command("copy-structure-missing", context_settings=CONTEXT_SETTINGS)
@click.argument(
    "path", type=click.Path(exists=True, file_okay=False, dir_okay=True, path_type=Path)
)
@click.argument(
    "path2",
    type=click.Path(exists=True, file_okay=False, dir_okay=True, path_type=Path),
)
@click.argument(
    "out_dir", type=click.Path(file_okay=False, dir_okay=True, path_type=Path)
)
@click.option(
    "--prefix",
    default="Обработано__",
    show_default=True,
    help="Префикс в именах файлов, который нужно игнорировать при сравнении. "
    "После снятия также убираются ведущие '_'/' ' и '-' сразу за префиксом.",
)
@click.option(
    "--prefix-side",
    type=click.Choice(["path", "path2", "both"]),
    default="path",
    show_default=True,
    help="На какой стороне снимать префикс из имён файлов (PATH, PATH2 или обеих). "
    "Если префикс в PATH2 — укажите 'path2'.",
)
@click.option(
    "--ext",
    multiple=True,
    default=[".xlsx", ".xls"],
    show_default=True,
    help="Какие расширения сравнивать/копировать (можно указать несколько раз).",
)
@click.option(
    "--equate-excel-ext/--strict-ext",
    default=True,
    show_default=True,
    help="Считать .xls и .xlsx эквивалентными при сравнении (копируется исходный файл как есть).",
)
@click.option(
    "--dest-naming",
    type=click.Choice(["path", "path2"]),
    default="path2",
    show_default=True,
    help="Как именовать копии в OUT_DIR: как в PATH (без префикса) или как в PATH2 (с префиксом, если он там используется).",
)
@click.option(
    "--prefix-joiner",
    default="__",
    show_default=True,
    help="Разделитель между префиксом и именем при формировании имён в стиле PATH2.",
)
@click.option(
    "--dry-run/--no-dry-run",
    default=False,
    show_default=True,
    help="Только показать, что будет скопировано, без фактического копирования.",
)
def copy_structure_missing(
    path: Path,
    path2: Path,
    out_dir: Path,
    prefix: str,
    prefix_side: str,
    ext: tuple[str, ...],
    equate_excel_ext: bool,
    dest_naming: str,
    prefix_joiner: str,
    dry_run: bool,
):
    """
    Находит файлы, которых нет в PATH2 (учитывая префикс/нормализацию), и копирует их из PATH в OUT_DIR,
    сохраняя относительную структуру. Имена назначения можно сделать в стиле PATH или PATH2.
    """
    import unicodedata

    exts = {e.lower() for e in ext}
    out_dir.mkdir(parents=True, exist_ok=True)

    def u_norm(s: str) -> str:
        return unicodedata.normalize("NFC", s)

    def strip_prefix(name: str, pref: str) -> str:
        name = u_norm(name)
        pref = u_norm(pref)
        if name.startswith(pref):
            rest = name[len(pref) :]
            # убираем ведущие подчёркивания/пробелы/дефисы
            rest = re.sub(r"^[_\s-]+", "", rest)
            return rest
        return name

    def normalize_basename(basename: str, apply_prefix_strip: bool) -> str:
        name = strip_prefix(basename, prefix) if apply_prefix_strip else basename
        name = u_norm(name)
        # Выравнивание расширений только для сравнения
        if equate_excel_ext:
            suff = Path(name).suffix.lower()
            if suff in {".xls", ".xlsx"}:
                name = Path(name).with_suffix(".xlsx").name
        return name

    def normalize_rel(rel: Path, side_tag: str) -> Path:
        """Нормализованный относительный путь для сравнения (снятие префикса и приведение расширения)."""
        apply = prefix_side in (side_tag, "both")
        norm_name = normalize_basename(rel.name, apply_prefix_strip=apply)
        return rel.with_name(norm_name)

    def add_prefix_for_path2(name: str) -> str:
        """Формирует имя в стиле PATH2 (с префиксом, если он используется на стороне PATH2)."""
        if prefix_side in ("path2", "both"):
            # не дублируем префикс
            if name.startswith(prefix):
                return name
            return (
                f"{prefix}{prefix_joiner}{name}" if prefix_joiner else f"{prefix}{name}"
            )
        return name

    # --- Индексация левой стороны (PATH): нормализованный относительный путь -> исходный путь и оригинальный rel ---
    left_map: dict[Path, tuple[Path, Path]] = {}
    collisions_left: dict[Path, list[Path]] = {}

    for p in path.rglob("*"):
        if not p.is_file() or p.name.startswith("~$") or p.suffix.lower() not in exts:
            continue
        rel = p.relative_to(path)
        rel_norm = normalize_rel(rel, "path")
        if rel_norm in left_map:
            collisions_left.setdefault(rel_norm, []).append(rel)
            # оставляем первый встретившийся, но предупреждаем
        else:
            left_map[rel_norm] = (p, rel)

    # --- Индексация правой стороны (PATH2) ---
    right_norm_set: set[Path] = set()
    collisions_right: dict[Path, list[Path]] = {}

    for q in path2.rglob("*"):
        if not q.is_file() or q.name.startswith("~$") or q.suffix.lower() not in exts:
            continue
        rel = q.relative_to(path2)
        rel_norm = normalize_rel(rel, "path2")
        if rel_norm in right_norm_set:
            collisions_right.setdefault(rel_norm, []).append(rel)
        right_norm_set.add(rel_norm)

    # --- Вычисляем недостающие ---
    missing_normals = sorted([nr for nr in left_map.keys() if nr not in right_norm_set])

    click.echo(
        f"Поиск недостающих файлов:\n"
        f"  FROM: {path}\n"
        f"  TO  : {path2}\n"
        f"  OUT : {out_dir}\n"
        f"  Префикс: '{prefix}' (side: {prefix_side}, joiner: '{prefix_joiner}')\n"
        f"  Эквивалентность .xls/.xlsx при сравнении: {'ON' if equate_excel_ext else 'OFF'}\n"
        f"  Стиль имён назначения: {dest_naming}\n"
    )

    if collisions_left:
        click.echo(
            "⚠️ Коллизии после нормализации на стороне PATH (несколько файлов сводятся к одному пути):"
        )
        for rel_norm, original_rels in collisions_left.items():
            click.echo(
                f"   • {rel_norm} <= " + "; ".join(str(x) for x in original_rels)
            )
        click.echo("")

    if collisions_right:
        click.echo("⚠️ Коллизии после нормализации на стороне PATH2:")
        for rel_norm, original_rels in collisions_right.items():
            click.echo(
                f"   • {rel_norm} <= " + "; ".join(str(x) for x in original_rels)
            )
        click.echo("")

    if not missing_normals:
        click.echo("✅ Недостающих файлов не найдено.")
        return

    # --- Копирование ---
    copied = 0
    errors = 0

    for rel_norm in tqdm(missing_normals, desc="Копирование недостающих"):
        src_abs, rel_orig = left_map[rel_norm]

        # Определяем относительный путь назначения
        if dest_naming == "path":
            # как в PATH (оригинальные имена)
            dest_rel = rel_orig
        else:
            # как в PATH2: берём НОРМАЛИЗОВАННЫЙ rel (без префикса) и при необходимости добавляем префикс к basename
            base_name = rel_norm.name
            base_name = add_prefix_for_path2(base_name)
            dest_rel = rel_norm.with_name(base_name)

        dst_abs = (out_dir / dest_rel).resolve()
        dst_abs.parent.mkdir(parents=True, exist_ok=True)

        if dry_run:
            click.echo(f"[DRY] {src_abs}  ->  {dst_abs}")
            continue

        try:
            shutil.copy2(src_abs, dst_abs)
            copied += 1
        except Exception as e:
            errors += 1
            click.echo(f"❌ Ошибка копирования {src_abs} -> {dst_abs}: {e}", err=True)

    if dry_run:
        click.echo(
            f"\nDRY-RUN: всего было бы скопировано файлов: {len(missing_normals)}"
        )
    else:
        click.echo(f"\nГотово. Скопировано файлов: {copied}. Ошибок: {errors}.")


if __name__ == "__main__":
    # data_operations.callback(
    #     Path(
    #         r"puid_2024_missing_110925/03_ОП_МЗ_03Н-134_г.Белореченск_-_п.Нижневеденеевский"
    #     ),
    #     "cd",  # format: "cd" или "xg"
    #     None,  # out_base: Path(...) или None
    #     False,  # overwrite
    #     "_processed",  # suffix
    # )
    main()
