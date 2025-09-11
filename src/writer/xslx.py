from pathlib import Path
import logging
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def write_to_predict_sheet(
    path: Path,
    initial_df: pd.DataFrame,
    predicted: pd.DataFrame,
    header_template: Path | None = None,
):
    """
    Создаёт файл '<parent>/Обработано__{stem}.xlsx':
      1) читает header.xlsx и берёт ЕГО ЯЧЕЙКИ как «шапку»,
      2) добавляет листы 'Исходные данные', 'Результаты', 'Объединенные данные',
      3) на каждый лист сначала пишет шапку, затем ТОЛЬКО ЗНАЧЕНИЯ DF (без заголовков), индекс -> первый столбец (datetime).
      В книге НЕТ отдельного листа из шаблона.
    """
    dst = path.parent / f"Обработано__{path.stem}.xlsx"
    template = Path("header.xlsx") if header_template is None else Path(header_template)

    # --- 0) Считываем значения из header.xlsx (если есть) ---
    header_values: list[list] = []
    if template.exists():
        tmpl_wb = load_workbook(template, data_only=True, read_only=True)
        try:
            tmpl_ws = tmpl_wb.worksheets[0]
            max_row, max_col = tmpl_ws.max_row, tmpl_ws.max_column
            for r in range(1, max_row + 1):
                header_values.append(
                    [tmpl_ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
                )
        finally:
            try:
                tmpl_wb.close()
            except Exception:
                pass

    # --- 1) Готовим чистую книгу (без листов из шаблона) ---
    wb = Workbook()
    wb.remove(wb.active)  # убираем дефолтный пустой лист

    # --- вспомогалки как у вас ---
    def _rows_only(df: pd.DataFrame) -> list[list]:
        """DF -> список строк: [index_value, *row_values], без заголовков."""
        idx = pd.to_datetime(df.index, errors="coerce")
        idx_vals = [
            (ts.to_pydatetime().replace(tzinfo=None) if pd.notna(ts) else None)
            for ts in idx
        ]
        arr = df.to_numpy(dtype=object)
        rows = []
        for i, row in enumerate(arr):
            cleaned = []
            for v in row:
                if pd.isna(v) or (isinstance(v, float) and not np.isfinite(v)):
                    cleaned.append(None)
                else:
                    try:
                        import numpy as _np

                        if isinstance(v, _np.generic):
                            v = v.item()
                    except Exception:
                        pass
                    cleaned.append(v)
            rows.append([idx_vals[i], *cleaned])
        return rows

    def _write_sheet_with_header(wb, sheet_name: str, data_rows: list[list]):
        ws = wb.create_sheet(title=sheet_name)
        # 1) сначала — строки из header.xlsx (если они есть)
        for r in header_values:
            ws.append(r)
        # 2) затем — данные
        for r in data_rows:
            ws.append(r)
        # формат столбца A как datetime
        a = get_column_letter(1)
        ws.column_dimensions[a].width = 20
        for cell in ws[a]:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"

    try:
        # 2) подготовка данных (без заголовков)
        num = predicted.apply(pd.to_numeric, errors="coerce").replace(
            [np.inf, -np.inf], np.nan
        )
        arr = num.to_numpy(dtype=float)

        info = np.iinfo(np.int64)
        # округляем (banker's rounding). Можно заменить на np.floor/np.ceil при желании
        arr = np.rint(arr)

        # чистим «плохие» значения: NaN/inf уже NaN, отбрасываем выход за диапазон int64
        bad = ~np.isfinite(arr) | (arr < info.min) | (arr > info.max)
        arr[bad] = np.nan

        rounded_df = pd.DataFrame(arr, index=num.index, columns=num.columns)
        pred_num = rounded_df.astype("Int64")  # теперь безопасно
        init_rows = _rows_only(initial_df)
        pred_rows = _rows_only(pred_num)
        comb_rows = _rows_only(pred_num.combine_first(initial_df))

        # 3) запись трёх листов (каждый лист начинается с header.xlsx)
        _write_sheet_with_header(wb, "Исходные данные", init_rows)
        _write_sheet_with_header(wb, "Результаты", pred_rows)
        _write_sheet_with_header(wb, "Объединенные данные", comb_rows)

        # 4) сохранить
        dst.parent.mkdir(parents=True, exist_ok=True)
        wb.save(dst)
        wb.close()
        logger.info(f"Excel сохранён: {dst}")
    except Exception as e:
        logger.error(f"Writing to Excel crashed: {e}")
